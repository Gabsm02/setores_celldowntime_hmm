from dotenv import load_dotenv
load_dotenv()

import io
import os
from typing import List

import urllib3
import pandas as pd
import openpyxl
from zipfile import ZipFile

# ==========================
# Configurações
# ==========================

URL_ZIP = os.getenv("URL_ZIP")
ALVO_NO_ZIP = os.getenv("ALVO_NO_ZIP")
PASTA_SAIDA = "./saida"
NOME_ARQUIVO_SAIDA = "Setores_Celldowntime_BA.xlsx"

# NOVA PLANILHA PARA MERGE (opcional)
NOVO_ARQUIVO = "./nova_planilha.xlsx"

COLUNAS_FIXAS = [
    "REGIONAL", "UF", "MUNICIPIO", "CN", "SITE",
    "TECNOLOGIA", "ERB", "SETOR"
]

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
HTTP = urllib3.PoolManager(cert_reqs="CERT_NONE")


# ==========================
# Utilitários
# ==========================
def baixar_zip(url: str) -> bytes:
    print("Baixando ZIP...")
    resp = HTTP.request("GET", url)
    if resp.status != 200:
        raise RuntimeError(f"Falha ao baixar ZIP. HTTP {resp.status}")
    return resp.data


def extrair_arquivo_do_zip(zip_bytes: bytes, alvo: str, destino: str) -> str:
    os.makedirs(destino, exist_ok=True)
    print("Abrindo ZIP...")
    with ZipFile(io.BytesIO(zip_bytes)) as zf:
        print("Arquivos encontrados:", zf.namelist())
        if alvo not in zf.namelist():
            raise FileNotFoundError(
                f"'{alvo}' não está dentro do ZIP. Disponíveis: {zf.namelist()}"
            )
        caminho = zf.extract(alvo, path=destino)
    print(f"Arquivo extraído para: {caminho}")
    return caminho


def detectar_aba_com_dados(caminho_xlsx: str) -> str:
    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if ws.max_row > 1 and ws.max_column > 1:
            print(f"Aba com dados detectada automaticamente: {sheet}")
            return sheet
    raise ValueError("Nenhuma aba com dados encontrada.")


def detectar_linha_header(caminho_xlsx: str, aba: str, tentar_primeiras: int = 10) -> int:
    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    ws = wb[aba]

    for i in range(1, tentar_primeiras + 1):
        values = [(str(c.value).strip() if c.value is not None else "") for c in ws[i]]
        score = sum(1 for v in values if v and not v.isdigit())
        if score >= 2:
            print(f"Linha de cabeçalho detectada: {i}")
            return i - 1

    print("Nenhum cabeçalho claro detectado. Usando 1ª linha.")
    return 0


def ler_excel(caminho_xlsx: str, aba: str, header_row: int) -> pd.DataFrame:
    df = pd.read_excel(
        caminho_xlsx,
        sheet_name=aba,
        header=header_row,
        engine="openpyxl",
    )
    print(f"Colunas detectadas ({len(df.columns)}): {list(df.columns)}")
    return df


def filtrar_bahia(df: pd.DataFrame) -> pd.DataFrame:
    if "UF" in df.columns:
        return df[df["UF"] == "BA"].copy()
    print("Sem coluna 'UF'. Nenhum filtro aplicado.")
    return df.copy()


def detectar_colunas_data(df: pd.DataFrame) -> List[str]:
    datas = []
    for col in df.columns:
        try:
            pd.to_datetime(col, dayfirst=True, errors="raise")
            datas.append(col)
        except Exception:
            pass
    return datas


def escolher_coluna_data_mais_recente(colunas_data: List[str]) -> str:
    if not colunas_data:
        raise ValueError("Nenhuma coluna de data encontrada.")
    return max(colunas_data, key=lambda x: pd.to_datetime(x, dayfirst=True))


# ==========================
# NOVAS FUNÇÕES
# ==========================
def filtrar_valores_maiores_500(df: pd.DataFrame, coluna_data: str) -> pd.DataFrame:
    print(f"Filtrando valores >= 500 na coluna: {coluna_data}")
    return df[df[coluna_data] >= 500].copy()


def adicionar_contagem_erb_e_agrupar(df: pd.DataFrame, coluna_data: str) -> pd.DataFrame:
    print("Gerando QTD_ERB e consolidando uma linha por ERB...")

    # conta quantas linhas existem para cada ERB
    df["QTD_ERB"] = df.groupby("ERB")["ERB"].transform("count")

    # consolida uma linha por ERB, pegando somente o primeiro valor da data
    df_grouped = df.groupby("ERB", as_index=False).agg(
        {
            coluna_data: "first",      # pega o primeiro valor da data
            "QTD_ERB": "max",          # qnt total de linhas da ERB
            "REGIONAL": "first",
            "UF": "first",
            "MUNICIPIO": "first",
            "CN": "first",
            "SITE": "first",
            "TECNOLOGIA": "first",
            "SETOR": "first"
        }
    )

    return df_grouped


def mesclar_com_nova_planilha(df: pd.DataFrame, caminho: str) -> pd.DataFrame:
    if not os.path.exists(caminho):
        print("Nenhuma nova planilha encontrada para merge.")
        return df

    print(f"Mergindo com nova planilha: {caminho}")
    df_novo = pd.read_excel(caminho, engine="openpyxl")

    if "ERB" not in df_novo.columns:
        print("Planilha nova não possui coluna ERB. Merge ignorado.")
        return df

    return df.merge(df_novo, on="ERB", how="left", suffixes=("", "_NOVO"))


def salvar_excel(df: pd.DataFrame, pasta: str, nome_arquivo: str) -> str:
    os.makedirs(pasta, exist_ok=True)
    caminho = os.path.join(pasta, nome_arquivo)
    df.to_excel(caminho, index=False)
    print(f"Arquivo salvo: {caminho}")
    return caminho


# ==========================
# Orquestração
# ==========================
def main():
    try:
        zip_bytes = baixar_zip(URL_ZIP)
        caminho_extraido = extrair_arquivo_do_zip(zip_bytes, ALVO_NO_ZIP, PASTA_SAIDA)

        aba = detectar_aba_com_dados(caminho_extraido)
        header_row = detectar_linha_header(caminho_extraido, aba)

        df = ler_excel(caminho_extraido, aba, header_row)
        df_ba = filtrar_bahia(df)

        col_datas = detectar_colunas_data(df_ba)
        col_data_mais_recente = escolher_coluna_data_mais_recente(col_datas)
        print("Coluna de data mais recente detectada:", col_data_mais_recente)

        df_ba = filtrar_valores_maiores_500(df_ba, col_data_mais_recente)

        df_ba = adicionar_contagem_erb_e_agrupar(df_ba, col_data_mais_recente)

        df_ba = mesclar_com_nova_planilha(df_ba, NOVO_ARQUIVO)

        salvar_excel(df_ba, PASTA_SAIDA, NOME_ARQUIVO_SAIDA)

        print("Concluído com sucesso!")

    except Exception as e:
        print(f"Erro: {e}")


if __name__ == "__main__":
    main()
